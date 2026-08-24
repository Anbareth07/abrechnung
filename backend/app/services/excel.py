"""Excel-Export der Jahresabrechnung (komplette Arbeitsmappe) mit openpyxl.

Enthält: Matrix (Kostenarten × Mieter) plus alle Ausgangsdaten des
Abrechnungsjahres (Stammdaten, Rechnungen, Strom, Wasser, Techem,
Vorauszahlungen) – als eigenständige, nachvollziehbare Abrechnungsdatei.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from .. import models
from ..models.enums import InvoiceKind
from . import strom as strom_service
from . import wasser as wasser_service
from .engine import _anteil_factor, compute_settlement
from .prorata import year_bounds

_CENTS = Decimal("0.01")

# Umlageschlüssel → Anzeige-Label (wie in der Abrechnungsansicht)
_ALLOC_LABELS = {
    "WF": "Wohnfläche",
    "NF": "Nutzfläche",
    "CONSUMPTION": "Verbrauch",
    "WOHNUNG": "Wohnung",
    "NONE": "—",
}

_TITLE_FONT = Font(bold=True, size=14)
_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="DEE2E6")
_SUM_FILL = PatternFill("solid", fgColor="E7F5FF")
_SALDO_FILL = PatternFill("solid", fgColor="FFF3BF")
_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center")
_RIGHT = Alignment(horizontal="right")


def _fmt_date(d: date) -> str:
    return d.strftime("%d.%m.%Y")


def _money(value) -> float:
    """Wert als auf 2 Nachkommastellen gerundeter float (für Excel-Zellen)."""
    if value is None:
        return 0.0
    return float(Decimal(str(value)).quantize(_CENTS))


def _num(value):
    """Wert als float ohne Rundung (z. B. Zählerstände)."""
    if value is None:
        return None
    return float(Decimal(str(value)))


def _write_table(ws, start_row, headers, rows, widths=None, money_cols=()):
    """Schreibt eine Tabelle mit Kopfzeile; gibt die letzte belegte Zeile zurück."""
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for r, row in enumerate(rows, start=start_row + 1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(r, c, value)
            cell.border = _BORDER
            if c in money_cols:
                cell.number_format = "#,##0.00"
                cell.alignment = _RIGHT
            elif isinstance(value, (int, float)):
                cell.alignment = _RIGHT
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return start_row + len(rows)


def _add_stammdaten_sheet(wb, session, result, property_id, year, ys, ye):
    """Objekt, Einheiten und im Jahr aktive Mieter."""
    prop = session.get(models.Property, property_id)
    ws = wb.create_sheet("Stammdaten")
    ws["A1"] = f"Stammdaten {year}"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = "Objekt, Einheiten und Mieter als Grundlage der Berechnung"
    ws["A2"].font = Font(italic=True, size=9)

    r = 4
    ws.cell(r, 1, "Objekt").font = _HEADER_FONT
    r += 1
    obj_rows = [
        ("Name", result.property_name),
        ("Straße", prop.street if prop else ""),
        ("PLZ", prop.zip_code if prop else ""),
        ("Ort", prop.city if prop else ""),
        ("Gesamtwohnfläche (m²)", _money(result.total_wf)),
        ("Gesamtnutzfläche (m²)", _money(result.total_nf)),
    ]
    if prop is not None and prop.wasser_versiegelte_flaeche is not None:
        obj_rows.append(("Versiegelte Fläche (m²)", _money(prop.wasser_versiegelte_flaeche)))
    for label, value in obj_rows:
        ws.cell(r, 1, label).border = _BORDER
        cell = ws.cell(r, 2, value)
        cell.border = _BORDER
        if isinstance(value, (int, float)):
            cell.number_format = "#,##0.00"
            cell.alignment = _RIGHT
        r += 1

    units = session.scalars(
        select(models.LeaseUnit)
        .where(models.LeaseUnit.property_id == property_id)
        .order_by(models.LeaseUnit.designation)
    ).all()

    r += 1
    ws.cell(r, 1, "Einheiten").font = _HEADER_FONT
    r += 1
    unit_rows = [
        (u.designation, _money(u.living_area), _money(u.extra_area), _money(u.utility_area))
        for u in units
    ]
    r = _write_table(
        ws, r, ["Bezeichnung", "Wohnfläche (m²)", "Garagenfläche (m²)", "Nutzfläche (m²)"],
        unit_rows, widths=[24, 18, 18, 18], money_cols=(2, 3, 4),
    ) + 1

    ws.cell(r, 1, "Mieter (im Abrechnungsjahr aktiv)").font = _HEADER_FONT
    r += 1
    tenant_rows = []
    for u in units:
        for t in u.tenants:
            occ_start = max(t.move_in, ys)
            occ_end = min(t.move_out, ye) if t.move_out else ye
            if occ_start > occ_end:
                continue
            days = (occ_end - occ_start).days + 1
            tenant_rows.append(
                (
                    t.name,
                    u.designation,
                    _fmt_date(t.move_in),
                    _fmt_date(t.move_out) if t.move_out else "",
                    _money(t.monthly_advance),
                    days,
                )
            )
    _write_table(
        ws, r, ["Name", "Wohnung", "Einzug", "Auszug", "Vorauszahlung (€/Monat)", "Tage im Jahr"],
        tenant_rows, widths=[24, 20, 12, 12, 22, 12], money_cols=(5,),
    )


def _add_rechnungen_sheet(wb, session, property_id, year, ys, ye):
    """Alle Rechnungen des Jahres inkl. Anrechnungsanteil und angerechnetem Betrag."""
    ws = wb.create_sheet("Rechnungen")
    ws["A1"] = f"Rechnungen {year}"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = "Eingabedaten (reale Beträge) und daraus angerechnete Beträge"
    ws["A2"].font = Font(italic=True, size=9)

    invoices = session.scalars(
        select(models.Invoice).where(models.Invoice.property_id == property_id)
    ).all()
    rows = []
    for inv in invoices:
        overlaps = (inv.valid_from is not None and inv.valid_from <= ye) or (
            inv.period_start <= ye and inv.period_end >= ys
        )
        if not overlaps:
            continue
        cat = inv.cost_category
        if (
            inv.kind == InvoiceKind.GRUNDSTEUER
            and inv.valid_from is not None
            and inv.annual_amount is not None
        ):
            real = inv.annual_amount
            period = f"{_fmt_date(inv.valid_from)} – (Bescheid)"
        else:
            real = sum((it.gross_amount for it in inv.items), Decimal("0"))
            period = f"{_fmt_date(inv.period_start)} – {_fmt_date(inv.period_end)}"
        faktor = _anteil_factor(inv)
        anteil_label = ""
        if inv.anteil_zaehler is not None and inv.anteil_nenner is not None:
            anteil_label = f"{int(Decimal(str(inv.anteil_zaehler)))}/{int(Decimal(str(inv.anteil_nenner)))}"
        comment = ""
        if inv.meta and "kommentar" in inv.meta:
            comment = str(inv.meta["kommentar"])
        rows.append(
            (
                cat.name if cat else "",
                inv.kind or "",
                period,
                inv.description or "",
                inv.invoice_number or "",
                inv.supplier or "",
                _money(real),
                anteil_label,
                _money(real * faktor),
                comment,
            )
        )
    rows.sort(key=lambda x: (x[0] or "", x[3] or ""))
    _write_table(
        ws, 4,
        ["Kostenstelle", "Art", "Zeitraum", "Titel", "Rechnungsnr.", "Lieferant",
         "Betrag (€)", "Anteil", "angerechnet (€)", "Kommentar"],
        rows,
        widths=[22, 18, 26, 28, 14, 20, 14, 12, 16, 26],
        money_cols=(7, 9),
    )


def _add_strom_sheet(wb, session, property_id, year, ys, ye):
    """Tarife, Zählerstände und Berechnung des Strom-Moduls."""
    ws = wb.create_sheet("Strom")
    ws["A1"] = f"Strom {year}"
    ws["A1"].font = _TITLE_FONT
    start_bound = date(year - 1, 12, 31)

    r = 3
    ws.cell(r, 1, "Tarife").font = _HEADER_FONT
    r += 1
    prices = session.scalars(
        select(models.StromPrice)
        .where(models.StromPrice.property_id == property_id)
        .order_by(models.StromPrice.kind, models.StromPrice.valid_from)
    ).all()
    price_rows = [
        (p.kind, _fmt_date(p.valid_from), _fmt_date(p.valid_to), _money(p.amount), _money(p.vat_rate))
        for p in prices
    ]
    r = _write_table(
        ws, r, ["Art", "Gültig von", "Gültig bis", "Betrag", "MwSt %"],
        price_rows, widths=[18, 12, 12, 12, 10], money_cols=(4, 5),
    ) + 1

    ws.cell(r, 1, "Zählerstände").font = _HEADER_FONT
    r += 1
    readings = session.scalars(
        select(models.StromReading)
        .where(models.StromReading.property_id == property_id)
        .order_by(models.StromReading.role, models.StromReading.reading_date)
    ).all()
    read_rows = [
        (
            x.role,
            _fmt_date(x.reading_date),
            _num(x.value),
            "ja" if x.vor_zaehlerwechsel else "",
            _num(x.neuer_zaehler_start) if x.vor_zaehlerwechsel else "",
        )
        for x in readings
        if start_bound <= x.reading_date <= ye
    ]
    r = _write_table(
        ws, r, ["Rolle", "Datum", "Stand (kWh)", "Zählerwechsel", "Neustart"],
        read_rows, widths=[18, 12, 14, 14, 12], money_cols=(5,),
    ) + 1

    try:
        res = strom_service.berechnung(session, property_id, ys, ye)
    except ValueError:
        res = None
    if res is not None and res.get("hauptzaehler") is not None:
        ws.cell(r, 1, "Berechnung").font = _HEADER_FONT
        r += 1
        hz = res["hauptzaehler"]
        calc_rows = [("Hauptzähler", None, _num(hz["consumption"]), None, None, None)]
        unter = res.get("unterzaehler")
        if unter is not None and Decimal(str(unter.get("consumption", 0))) > 0:
            calc_rows.append(("− Unterzähler (Heizstrom)", None, _num(unter["consumption"]), None, None, None))
            calc_rows.append(("= Nettoverbrauch", None, _num(res.get("netto_verbrauch", 0)), None, None, None))
        for pos in res["positionen"]:
            calc_rows.append(
                (
                    pos["art"],
                    _money(pos["satz"]),
                    _num(pos["menge"]),
                    _money(pos["netto"]),
                    _money(pos["vat"]),
                    _money(pos["brutto"]),
                )
            )
        _write_table(
            ws, r,
            ["Position", "Satz (€)", "Menge", "Netto (€)", "MwSt (€)", "Brutto (€)"],
            calc_rows, widths=[24, 12, 12, 12, 12, 12], money_cols=(2, 4, 5, 6),
        )


def _add_wasser_sheet(wb, session, property_id, year, ys, ye):
    """Tarife, Haupt-/Wohnungszählerstände und Berechnung des Wasser-Moduls."""
    ws = wb.create_sheet("Wasser")
    ws["A1"] = f"Wasser {year}"
    ws["A1"].font = _TITLE_FONT
    start_bound = date(year - 1, 12, 31)

    r = 3
    ws.cell(r, 1, "Tarife").font = _HEADER_FONT
    r += 1
    prices = session.scalars(
        select(models.WasserPrice)
        .where(models.WasserPrice.property_id == property_id)
        .order_by(models.WasserPrice.kind, models.WasserPrice.valid_from)
    ).all()
    price_rows = [
        (p.kind, _fmt_date(p.valid_from), _fmt_date(p.valid_to), _money(p.amount), _money(p.vat_rate))
        for p in prices
    ]
    r = _write_table(
        ws, r, ["Art", "Gültig von", "Gültig bis", "Betrag", "MwSt %"],
        price_rows, widths=[18, 12, 12, 12, 10], money_cols=(4, 5),
    ) + 1

    ws.cell(r, 1, "Hauptzählerstände").font = _HEADER_FONT
    r += 1
    readings = session.scalars(
        select(models.WasserReading)
        .where(models.WasserReading.property_id == property_id)
        .order_by(models.WasserReading.reading_date)
    ).all()
    read_rows = [
        (
            _fmt_date(x.reading_date),
            _num(x.value),
            "ja" if x.vor_zaehlerwechsel else "",
            _num(x.neuer_zaehler_start) if x.vor_zaehlerwechsel else "",
        )
        for x in readings
        if start_bound <= x.reading_date <= ye
    ]
    r = _write_table(
        ws, r, ["Datum", "Stand (m³)", "Zählerwechsel", "Neustart"],
        read_rows, widths=[12, 14, 14, 12], money_cols=(4,),
    ) + 1

    units = session.scalars(
        select(models.LeaseUnit).where(models.LeaseUnit.property_id == property_id)
    ).all()
    unit_ids = [u.id for u in units]
    meters = session.scalars(
        select(models.Meter)
        .where(models.Meter.lease_unit_id.in_(unit_ids))
        .order_by(models.Meter.name, models.Meter.meter_type)
    ).all()
    if meters:
        ws.cell(r, 1, "Wohnungszähler").font = _HEADER_FONT
        r += 1
        meter_rows = []
        for m in meters:
            for rd in m.readings:
                if start_bound <= rd.reading_date <= ye:
                    meter_rows.append(
                        (
                            m.name,
                            m.lease_unit.designation if m.lease_unit else "",
                            m.meter_type.value,
                            _fmt_date(rd.reading_date),
                            _num(rd.value),
                            "ja" if rd.vor_zaehlerwechsel else "",
                            _num(rd.neuer_zaehler_start) if rd.vor_zaehlerwechsel else "",
                        )
                    )
        r = _write_table(
            ws, r,
            ["Zähler", "Wohnung", "Art", "Datum", "Stand (m³)", "Zählerwechsel", "Neustart"],
            meter_rows, widths=[24, 20, 18, 12, 12, 14, 12], money_cols=(7,),
        ) + 1

    try:
        res = wasser_service.berechnung(session, property_id, ys, ye)
    except ValueError:
        res = None
    if res is not None:
        ws.cell(r, 1, "Berechnung").font = _HEADER_FONT
        r += 1
        plan = res.get("plan")
        calc_rows = []
        if plan == "B" and res.get("hauptzaehler"):
            calc_rows.append(("Hauptzähler", None, _num(res["hauptzaehler"]["consumption"]), None, None, None))
        elif plan == "A":
            calc_rows.append(("Wohnungsverbrauch", None, _num(res["verbrauch"]), None, None, None))
        for pos in res["positionen"]:
            calc_rows.append(
                (
                    pos["art"],
                    _money(pos["satz"]),
                    _num(pos["menge"]),
                    _money(pos["netto"]),
                    _money(pos["vat"]),
                    _money(pos["brutto"]),
                )
            )
        _write_table(
            ws, r,
            ["Position", "Satz (€)", "Menge", "Netto (€)", "MwSt (€)", "Brutto (€)"],
            calc_rows, widths=[24, 12, 12, 12, 12, 12], money_cols=(2, 4, 5, 6),
        )


def _add_techem_sheet(wb, session, property_id, year, ys, ye):
    """Heizkostenblätter (Techem), die das Abrechnungsjahr berühren."""
    ws = wb.create_sheet("Techem")
    ws["A1"] = f"Techem {year}"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = "Heizkostenblätter (fließen nicht in die Mieterabrechnung ein)"
    ws["A2"].font = Font(italic=True, size=9)

    records = session.scalars(
        select(models.TechemRecord)
        .where(models.TechemRecord.property_id == property_id)
        .order_by(models.TechemRecord.von)
    ).all()
    rows = [
        (
            _fmt_date(x.von),
            _fmt_date(x.bis),
            _num(x.gas_kwh),
            _money(x.gas_cost),
            _money(x.maintenance_cost),
            _money(x.chimney_cost),
            x.notes or "",
        )
        for x in records
        if not (x.bis < ys or x.von > ye)
    ]
    _write_table(
        ws, 4,
        ["Heizperiode von", "bis", "Gas (kWh)", "Gaskosten (€)", "Wartung (€)", "Kamin (€)", "Notizen"],
        rows, widths=[16, 12, 12, 14, 14, 12, 40], money_cols=(4, 5, 6),
    )


def _add_vorauszahlungen_sheet(wb, session, property_id, year, ys, ye):
    """Vorauszahlungen der im Jahr aktiven Mieter."""
    ws = wb.create_sheet("Vorauszahlungen")
    ws["A1"] = f"Vorauszahlungen {year}"
    ws["A1"].font = _TITLE_FONT

    units = session.scalars(
        select(models.LeaseUnit).where(models.LeaseUnit.property_id == property_id)
    ).all()
    rows = []
    for u in units:
        for t in u.tenants:
            occ_start = max(t.move_in, ys)
            occ_end = min(t.move_out, ye) if t.move_out else ye
            if occ_start > occ_end:
                continue
            payments = list(t.advance_payments)
            if payments:
                for ap in payments:
                    rows.append((t.name, u.designation, _fmt_date(ap.valid_from), _money(ap.amount)))
            else:
                rows.append((t.name, u.designation, "—", _money(t.monthly_advance)))
    _write_table(
        ws, 3, ["Mieter", "Wohnung", "Gültig ab", "Betrag (€/Monat)"],
        rows, widths=[24, 20, 12, 18], money_cols=(4,),
    )


def generate_settlement_excel(session, property_id: int, year: int) -> bytes:
    """Erzeugt die komplette xlsx-Arbeitsmappe für ein Objekt und Jahr."""
    result = compute_settlement(session, property_id, year)
    ys, ye = year_bounds(year)

    # Zeilen: alle konfigurierten Kostenstellen in Reihenfolge, danach zusätzliche
    # Detail-Codes (z. B. automatisch je Rechnungsart angelegte WOHNUNG-Kosten).
    rows: list[dict] = []  # {code, name, year_cost, basis}
    seen: set[str] = set()
    for cl in result.category_lines:
        rows.append(
            {
                "code": cl.code,
                "name": cl.name,
                "year_cost": cl.year_cost,
                "basis": _ALLOC_LABELS.get(cl.allocation_key, cl.allocation_key),
            }
        )
        seen.add(cl.code)
    for t in result.tenant_lines:
        for d in t.details:
            if d.code not in seen:
                rows.append(
                    {"code": d.code, "name": d.name, "year_cost": d.year_cost, "basis": d.basis_label}
                )
                seen.add(d.code)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Abrechnung {year}"

    # Titelblock
    ws["A1"] = f"Nebenkostenabrechnung {year}"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Objekt: {result.property_name}"
    ws["A3"] = f"Zeitraum: {_fmt_date(ys)} – {_fmt_date(ye)}"
    ws["A4"] = (
        f"Gesamtwohnfläche: {_money(result.total_wf):,.2f} m²  ·  "
        f"Gesamtnutzfläche: {_money(result.total_nf):,.2f} m²"
    )

    n_tenants = len(result.tenant_lines)
    header_row = 6
    first_data_row = header_row + 1

    # Kopfzeile: Kostenart | Verteilung | Gesamtkosten (€) | Mieter …
    headers = ["Kostenart", "Verteilung", "Gesamtkosten (€)"]
    for t in result.tenant_lines:
        headers.append(t.name if not t.designation else f"{t.name} · {t.designation}")
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Datenzeilen
    for r, row in enumerate(rows):
        excel_row = first_data_row + r
        name_cell = ws.cell(excel_row, 1, row["name"])
        name_cell.border = _BORDER
        ws.cell(excel_row, 2, row["basis"]).border = _BORDER
        cost = ws.cell(excel_row, 3, _money(row["year_cost"]))
        cost.number_format = "#,##0.00"
        cost.border = _BORDER
        cost.alignment = _RIGHT
        for i, t in enumerate(result.tenant_lines):
            amount = sum((d.amount for d in t.details if d.code == row["code"]), Decimal("0"))
            cell = ws.cell(excel_row, 4 + i, _money(amount))
            cell.number_format = "#,##0.00"
            cell.border = _BORDER
            cell.alignment = _RIGHT

    last_data_row = first_data_row + len(rows) - 1 if rows else first_data_row - 1

    # Summenzeilen: Summe Kosten, Vorauszahlung, Saldo
    def _write_summary_row(excel_row: int, label: str, values: list[Decimal], fill, bold: bool):
        cell = ws.cell(excel_row, 1, label)
        cell.font = _HEADER_FONT
        cell.fill = fill
        cell.border = _BORDER
        for i, v in enumerate(values):
            c = ws.cell(excel_row, 4 + i, _money(v))
            c.number_format = "#,##0.00"
            c.font = _HEADER_FONT if bold else None
            c.fill = fill
            c.border = _BORDER
            c.alignment = _RIGHT

    total_row = last_data_row + 1
    sum_costs = sum((row["year_cost"] for row in rows), Decimal("0"))
    ws.cell(total_row, 3, _money(sum_costs)).number_format = "#,##0.00"
    ws.cell(total_row, 3).fill = _SUM_FILL
    ws.cell(total_row, 3).border = _BORDER
    _write_summary_row(
        total_row,
        "Summe Kosten",
        [t.total_costs for t in result.tenant_lines],
        _SUM_FILL,
        True,
    )
    _write_summary_row(
        total_row + 1,
        "Vorauszahlung",
        [t.advance_total for t in result.tenant_lines],
        PatternFill(),
        False,
    )
    _write_summary_row(
        total_row + 2,
        "Saldo (Nachzahlung/Guthaben)",
        [t.saldo for t in result.tenant_lines],
        _SALDO_FILL,
        True,
    )

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    for i in range(n_tenants):
        ws.column_dimensions[get_column_letter(4 + i)].width = 18

    # Kopfzeile + erste 3 Spalten fixieren, Filter aktivieren
    ws.freeze_panes = "D7"
    if last_data_row >= first_data_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(3 + n_tenants)}{last_data_row}"

    # --- Ausgangsdaten des Abrechnungsjahres als weitere Blätter ----------------
    _add_stammdaten_sheet(wb, session, result, property_id, year, ys, ye)
    _add_rechnungen_sheet(wb, session, property_id, year, ys, ye)
    _add_strom_sheet(wb, session, property_id, year, ys, ye)
    _add_wasser_sheet(wb, session, property_id, year, ys, ye)
    _add_techem_sheet(wb, session, property_id, year, ys, ye)
    _add_vorauszahlungen_sheet(wb, session, property_id, year, ys, ye)

    # Hinweise als eigenes Blatt
    if result.warnings:
        ww = wb.create_sheet("Hinweise")
        ww["A1"] = "Hinweise zur Abrechnung"
        ww["A1"].font = _HEADER_FONT
        for i, w in enumerate(result.warnings):
            ww.cell(2 + i, 1, w)
        ww.column_dimensions["A"].width = 100

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
