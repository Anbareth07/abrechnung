"""Tests für komplexe Edge Cases der Jahresabrechnung.

Ergänzt die vorhandene Abdeckung um:
- Mieterwechsel (Vormieter + Nachmieter) in derselben Wohnung
- Zählerwechsel am Hauptzähler (Wasser-/Strom-Modul)
- Mehrere Rechnungen derselben Kostenstelle (normal + wiederkehrend kombiniert)
- Abdeckungs-Warnungen (Jahresanfang/-ende fehlt)
- Anrechnungsanteil auf wohneinheitenbezogene Rechnungen + Excel-Blatt
"""

from datetime import date
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest

from app import models
from app.models.enums import AllocationKey, InvoiceKind
from app.services import engine as engine_mod
from app.services import strom as strom_service
from app.services import wasser as wasser_service
from app.services.excel import generate_settlement_excel
from tests import helpers


# --- 1) Mieterwechsel (Vormieter + Nachmieter) ---------------------------------

def test_mieterwechsel_vormieter_nachmieter(session):
    """Klassischer Wechsel: Vormieter zieht 30.06. aus, Nachmieter 01.07. ein.

    Zeitfaktor, Vorauszahlung und Saldo müssen anteilig aufgeteilt sein,
    zusammen genau das Jahr (365 Tage) ergeben.
    """
    prop = helpers.make_property(session, "Testobjekt")
    grund = helpers.make_category(session, prop, "grundsteuer", "Grundsteuer", AllocationKey.WF)
    helpers.make_config(session, prop, grund, AllocationKey.WF, 1)
    u1 = helpers.make_unit(session, prop, "Wohnung 1", "50.0", "0.0")
    helpers.make_tenant(session, u1, "Vormieter", date(2020, 1, 1), "100.00", move_out=date(2026, 6, 30))
    helpers.make_tenant(session, u1, "Nachmieter", date(2026, 7, 1), "100.00")
    helpers.make_invoice(
        session, prop, grund, date(2026, 1, 1), date(2026, 12, 31),
        [(date(2026, 1, 1), date(2026, 12, 31), "1200.00")],
    )
    session.commit()

    result = engine_mod.compute_settlement(session, prop.id, 2026)
    vor = next(ln for ln in result.tenant_lines if ln.name == "Vormieter")
    nach = next(ln for ln in result.tenant_lines if ln.name == "Nachmieter")

    assert vor.tenant_days == 181  # 01.01.–30.06.
    assert nach.tenant_days == 184  # 01.07.–31.12.
    assert vor.tenant_days + nach.tenant_days == 365

    f_vor = Decimal(181) / Decimal(365)
    f_nach = Decimal(184) / Decimal(365)
    assert vor.time_factor == f_vor
    assert nach.time_factor == f_nach

    # WF 50/50 → voller Jahresbetrag × Zeitfaktor
    assert vor.breakdown["grundsteuer"] == Decimal("1200") * f_vor
    assert nach.breakdown["grundsteuer"] == Decimal("1200") * f_nach

    # Vorauszahlung: je 6 volle Monate
    assert vor.advance_total == Decimal("600.00")
    assert nach.advance_total == Decimal("600.00")
    # Saldo wird im Engine NICHT gerundet → ohne money() vergleichen
    assert vor.saldo == Decimal("1200") * f_vor - Decimal("600")
    assert nach.saldo == Decimal("1200") * f_nach - Decimal("600")


# --- 2) Zählerwechsel am Hauptzähler (Module) -----------------------------------

def _wechsel_lesung(model, pid, d, value, role=None):
    kwargs = {"property_id": pid, "reading_date": d, "value": Decimal(value)}
    if role is not None:
        kwargs["role"] = role
    r = model(**kwargs)
    r.vor_zaehlerwechsel = True
    r.neuer_zaehler_start = Decimal("0")
    return r


def test_zaehlerwechsel_hauptzaehler_wasser(session):
    """Wasser-Hauptzähler: letzter Stand alter Zähler (1000) + Start neuer Zähler (0).

    Neuer Zähler zeigt am 31.12. 42 → Verbrauch = 42 m³ (nicht negativ/fehlend).
    """
    prop = helpers.make_property(session, "Testobjekt")
    session.add(models.WasserPrice(
        property_id=prop.id, kind="TRINKWASSER", valid_from=date(2025, 1, 1),
        valid_to=date(2025, 12, 31), amount=Decimal("2.00"), vat_rate=Decimal("7.00"),
    ))
    session.add(_wechsel_lesung(models.WasserReading, prop.id, date(2025, 1, 1), "1000"))
    session.add(models.WasserReading(property_id=prop.id, reading_date=date(2025, 12, 31), value=Decimal("42")))
    session.commit()

    res = wasser_service.berechnung(session, prop.id, date(2025, 1, 1), date(2025, 12, 31))
    assert res["hauptzaehler"]["consumption"] == pytest.approx(42.0)
    trink = next(x for x in res["positionen"] if x["art"] == "TRINKWASSER")
    assert trink["netto"] == pytest.approx(84.0)


def test_zaehlerwechsel_hauptzaehler_strom(session):
    """Strom-Hauptzähler: Zählerwechsel → Verbrauch = 42 kWh statt Differenz 1000→42."""
    prop = helpers.make_property(session, "Testobjekt")
    session.add(models.StromPrice(
        property_id=prop.id, kind="ARBEITSPREIS", valid_from=date(2025, 1, 1),
        valid_to=date(2025, 12, 31), amount=Decimal("0.30"), vat_rate=Decimal("19.00"),
    ))
    session.add(_wechsel_lesung(models.StromReading, prop.id, date(2025, 1, 1), "1000", role="HAUPTZAEHLER"))
    session.add(models.StromReading(property_id=prop.id, role="HAUPTZAEHLER", reading_date=date(2025, 12, 31), value=Decimal("42")))
    session.commit()

    res = strom_service.berechnung(session, prop.id, date(2025, 1, 1), date(2025, 12, 31))
    assert res["hauptzaehler"]["consumption"] == pytest.approx(42.0)
    ap = next(x for x in res["positionen"] if x["art"] == "ARBEITSPREIS")
    assert ap["netto"] == pytest.approx(12.6)


# --- 3) Mehrere Rechnungen derselben Kostenstelle ------------------------------

def test_mehrere_normale_rechnungen_gleiche_kostenstelle(session):
    """Zwei normale Rechnungen derselben Kostenstelle werden summiert."""
    prop = helpers.make_property(session, "Testobjekt")
    grund = helpers.make_category(session, prop, "grundsteuer", "Grundsteuer", AllocationKey.WF)
    helpers.make_config(session, prop, grund, AllocationKey.WF, 1)
    u1 = helpers.make_unit(session, prop, "Wohnung 1", "50.0")
    helpers.make_tenant(session, u1, "Mieter A", date(2020, 1, 1), "100.00")
    helpers.make_invoice(
        session, prop, grund, date(2026, 1, 1), date(2026, 12, 31),
        [(date(2026, 1, 1), date(2026, 12, 31), "500.00")],
    )
    helpers.make_invoice(
        session, prop, grund, date(2026, 1, 1), date(2026, 12, 31),
        [(date(2026, 1, 1), date(2026, 12, 31), "300.00")],
    )
    session.commit()

    result = engine_mod.compute_settlement(session, prop.id, 2026)
    line = next(cl for cl in result.category_lines if cl.code == "grundsteuer")
    assert line.year_cost == Decimal("800.00")
    assert result.tenant_lines[0].breakdown["grundsteuer"] == Decimal("800.00")


def test_rechnung_wiederkehrend_plus_nachbescheid(session):
    """Wiederkehrende Grundsteuer (1200) + normale Nachbescheid-Rechnung (500) derselben Kostenstelle."""
    prop = helpers.make_property(session, "Testobjekt")
    grund = helpers.make_category(session, prop, "grundsteuer", "Grundsteuer", AllocationKey.WF)
    helpers.make_config(session, prop, grund, AllocationKey.WF, 1)
    u1 = helpers.make_unit(session, prop, "Wohnung 1", "50.0")
    helpers.make_tenant(session, u1, "Mieter A", date(2020, 1, 1), "100.00")

    inv = helpers.make_invoice(session, prop, grund, date(2026, 1, 1), date(2026, 12, 31), [])
    inv.kind = InvoiceKind.GRUNDSTEUER.value
    inv.valid_from = date(2020, 1, 1)
    inv.annual_amount = Decimal("1200.00")
    helpers.make_invoice(
        session, prop, grund, date(2026, 1, 1), date(2026, 12, 31),
        [(date(2026, 1, 1), date(2026, 12, 31), "500.00")],
    )
    session.commit()

    result = engine_mod.compute_settlement(session, prop.id, 2026)
    line = next(cl for cl in result.category_lines if cl.code == "grundsteuer")
    assert line.year_cost == Decimal("1700.00")


# --- 4) Abdeckungs-Warnungen (Jahresanfang/-ende fehlt) -------------------------

def test_abdeckungswarnung_wasser(session):
    """Nur ein Hauptzähler-Stand unterjährig → Warnungen für Jahresanfang UND -ende."""
    prop = helpers.make_property(session, "Testobjekt")
    session.add(models.WasserPrice(
        property_id=prop.id, kind="TRINKWASSER", valid_from=date(2025, 1, 1),
        valid_to=date(2025, 12, 31), amount=Decimal("2.00"), vat_rate=Decimal("7.00"),
    ))
    session.add(models.WasserReading(property_id=prop.id, reading_date=date(2025, 7, 1), value=Decimal("500")))
    session.commit()

    res = wasser_service.berechnung(session, prop.id, date(2025, 1, 1), date(2025, 12, 31))
    assert any("Jahresanfang" in w for w in res["warnings"])
    assert any("Jahresende" in w for w in res["warnings"])


def test_abdeckungswarnung_strom(session):
    """Strom-Hauptzähler-Stände decken den Zeitraum nicht ab → Warnungen."""
    prop = helpers.make_property(session, "Testobjekt")
    session.add(models.StromPrice(
        property_id=prop.id, kind="ARBEITSPREIS", valid_from=date(2025, 1, 1),
        valid_to=date(2025, 12, 31), amount=Decimal("0.30"), vat_rate=Decimal("19.00"),
    ))
    session.add(models.StromReading(property_id=prop.id, role="HAUPTZAEHLER", reading_date=date(2025, 7, 1), value=Decimal("500")))
    session.commit()

    res = strom_service.berechnung(session, prop.id, date(2025, 1, 1), date(2025, 12, 31))
    assert any("Jahresanfang" in w for w in res["warnings"])
    assert any("Jahresende" in w for w in res["warnings"])


def test_abdeckungswarnung_erreicht_abrechnung(session):
    """Die Wasser-Abdeckungswarnungen erscheinen in der Abrechnung (result.warnings)."""
    prop = helpers.make_property(session, "Testobjekt")
    grund = helpers.make_category(session, prop, "grundsteuer", "Grundsteuer", AllocationKey.WF)
    helpers.make_config(session, prop, grund, AllocationKey.WF, 1)
    u1 = helpers.make_unit(session, prop, "Wohnung 1", "50.0")
    helpers.make_tenant(session, u1, "Mieter A", date(2020, 1, 1), "100.00")
    session.add(models.WasserPrice(
        property_id=prop.id, kind="TRINKWASSER", valid_from=date(2025, 1, 1),
        valid_to=date(2025, 12, 31), amount=Decimal("2.00"), vat_rate=Decimal("7.00"),
    ))
    session.add(models.WasserReading(property_id=prop.id, reading_date=date(2025, 7, 1), value=Decimal("500")))
    session.commit()

    result = engine_mod.compute_settlement(session, prop.id, 2025)
    assert any("Wasser Hauptzähler" in w and "Jahresanfang" in w for w in result.warnings)
    assert any("Wasser Hauptzähler" in w and "Jahresende" in w for w in result.warnings)


# --- 5) Anrechnungsanteil auf WOHNUNG-Rechnung + Excel --------------------------

def test_anteil_wohneinheiten_rechnung(session):
    """Anrechnungsanteil wirkt auch auf wohneinheitenbezogene (WOHNUNG) Rechnungen."""
    prop = helpers.make_property(session, "Testobjekt")
    wohn = helpers.make_category(session, prop, "wohnungskosten", "Wohnungskosten", AllocationKey.WOHNUNG)
    helpers.make_config(session, prop, wohn, AllocationKey.WOHNUNG, 1)
    u1 = helpers.make_unit(session, prop, "Wohnung 1", "80.0")
    u2 = helpers.make_unit(session, prop, "Wohnung 2", "40.0")
    helpers.make_tenant(session, u1, "Mieter A", date(2020, 1, 1), "100.00")
    helpers.make_tenant(session, u2, "Mieter B", date(2020, 1, 1), "100.00")

    inv = helpers.make_invoice(
        session, prop, wohn, date(2026, 1, 1), date(2026, 12, 31),
        [(date(2026, 1, 1), date(2026, 12, 31), "1200.00")],
    )
    inv.lease_unit_id = u1.id
    inv.anteil_zaehler = Decimal("3")
    inv.anteil_nenner = Decimal("4")
    session.commit()

    result = engine_mod.compute_settlement(session, prop.id, 2026)
    a = next(ln for ln in result.tenant_lines if ln.name == "Mieter A")
    b = next(ln for ln in result.tenant_lines if ln.name == "Mieter B")
    assert a.breakdown["wohnungskosten"] == Decimal("900.00")  # 1200 × 3/4
    assert b.breakdown.get("wohnungskosten", Decimal("0")) == Decimal("0.00")


def test_excel_rechnungen_blatt_zeigt_anteil(session):
    """Das Excel-Blatt 'Rechnungen' zeigt Betrag, Anteil und angerechneten Betrag."""
    prop = helpers.make_property(session, "Testobjekt")
    grund = helpers.make_category(session, prop, "grundsteuer", "Grundsteuer", AllocationKey.WF)
    helpers.make_config(session, prop, grund, AllocationKey.WF, 1)
    u1 = helpers.make_unit(session, prop, "Wohnung 1", "50.0")
    helpers.make_tenant(session, u1, "Mieter A", date(2020, 1, 1), "100.00")

    inv = helpers.make_invoice(
        session, prop, grund, date(2026, 1, 1), date(2026, 12, 31),
        [(date(2026, 1, 1), date(2026, 12, 31), "1200.00")],
    )
    inv.description = "Testrechnung"
    inv.anteil_zaehler = Decimal("3")
    inv.anteil_nenner = Decimal("4")
    session.commit()

    data = generate_settlement_excel(session, prop.id, 2026)
    wb = openpyxl.load_workbook(BytesIO(data))
    assert "Rechnungen" in wb.sheetnames
    ws = wb["Rechnungen"]

    headers = [ws.cell(4, c).value for c in range(1, 11)]
    assert headers[6] == "Betrag (€)"
    assert headers[7] == "Anteil"
    assert headers[8] == "angerechnet (€)"

    for r in range(5, 30):
        if ws.cell(r, 4).value == "Testrechnung":
            assert ws.cell(r, 7).value == 1200.0
            assert ws.cell(r, 8).value == "3/4"
            assert ws.cell(r, 9).value == 900.0
            break
    else:
        raise AssertionError("Rechnung 'Testrechnung' nicht im Blatt 'Rechnungen' gefunden")
