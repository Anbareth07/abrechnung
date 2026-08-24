"""Tests für den Excel-Export der Jahresabrechnung."""

from io import BytesIO

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.db import get_db
from app.main import app
from app.models import Base
from app.services.excel import generate_settlement_excel
from tests.test_water import _build_objekt1
from tests.test_api import _seed_objekt1


@pytest.fixture
def client():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    Base.metadata.create_all(engine)
    TestingSession = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)

    def override_get_db():
        db = TestingSession()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()
    Base.metadata.drop_all(engine)


def test_excel_generation(session):
    """Service erzeugt eine gültige xlsx-Matrix (Kostenarten × Mieter)."""
    prop = _build_objekt1(session)

    data = generate_settlement_excel(session, prop.id, 2026)
    assert data[:2] == b"PK"  # zip/xlsx-Signatur

    wb = openpyxl.load_workbook(BytesIO(data))
    assert "Abrechnung 2026" in wb.sheetnames
    ws = wb["Abrechnung 2026"]

    assert ws["A1"].value == "Nebenkostenabrechnung 2026"
    assert ws["A2"].value == "Objekt: Objekt 1"

    # Kopfzeile: Kostenart | Verteilung | Gesamtkosten | 3 Mieter
    assert ws["A6"].value == "Kostenart"
    assert ws["B6"].value == "Verteilung"
    assert ws["C6"].value == "Gesamtkosten (€)"
    assert ws["D6"].value == "Mieter A · Wohnung 1"
    assert ws["E6"].value == "Mieter B · Wohnung 2"
    assert ws["F6"].value == "Mieter C · Wohnung 3"

    # Alle Kostenstellen als Zeilen vorhanden
    names = [ws.cell(7 + r, 1).value for r in range(20) if ws.cell(7 + r, 1).value]
    assert "Trinkwasser" in names
    assert "Schmutzwasser" in names
    assert "Grundsteuer" in names

    # Grundsteuer (NF): Verteilung "Nutzfläche", Gesamtkosten 2180
    grund_row = next(r for r in range(7, 40) if ws.cell(r, 1).value == "Grundsteuer")
    assert ws.cell(grund_row, 2).value == "Nutzfläche"
    assert ws.cell(grund_row, 3).value == 2180.0

    # Summenzeilen vorhanden
    labels = [ws.cell(r, 1).value for r in range(7, 40)]
    assert "Summe Kosten" in labels
    assert "Vorauszahlung" in labels
    assert "Saldo (Nachzahlung/Guthaben)" in labels


def test_excel_saldo_matches_api(session):
    """Saldo-Zeile im Excel entspricht der API-Berechnung (Kosten − Vorauszahlung)."""
    prop = _build_objekt1(session)
    data = generate_settlement_excel(session, prop.id, 2026)
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb["Abrechnung 2026"]

    # Vergleich mit compute_settlement
    from app.services.engine import compute_settlement

    result = compute_settlement(session, prop.id, 2026)
    labels = {}
    for r in range(7, 40):
        lab = ws.cell(r, 1).value
        if lab:
            labels[lab] = r
    for i, t in enumerate(result.tenant_lines):
        col = 4 + i
        assert abs(ws.cell(labels["Saldo (Nachzahlung/Guthaben)"], col).value - float(t.saldo)) < 0.01
        assert (
            abs(ws.cell(labels["Vorauszahlung"], col).value - float(t.advance_total)) < 0.01
        )
        assert abs(ws.cell(labels["Summe Kosten"], col).value - float(t.total_costs)) < 0.01


def test_excel_endpoint(client):
    """API-Endpoint liefert eine gültige xlsx-Datei mit korrekten Headern."""
    pid = _seed_objekt1(client)

    resp = client.get(f"/settlements/{pid}/2026/export.xlsx")
    assert resp.status_code == 200
    assert (
        resp.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "Abrechnung_2026_Objekt_1.xlsx" in resp.headers["content-disposition"]

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    assert "Abrechnung 2026" in wb.sheetnames
    ws = wb["Abrechnung 2026"]
    assert ws["A1"].value == "Nebenkostenabrechnung 2026"
    assert ws["D6"].value == "Mieter A · Wohnung 1"
