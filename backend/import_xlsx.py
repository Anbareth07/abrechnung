"""Import Zählerstände (Objekt 1) aus der Excel-Referenz in die DB.

Extrahiert aus allen Jahres-Blättern:
- Einheiten/Mieter (Name, Wohnfläche, Vorauszahlung)
- Wohnungs-Wasserzähler + Waschmaschinen-Zähler (Start/Ende)
- Garten Nord/Süd (Start/Ende)
und legt sie als Zählerstände (MeterReading) in der DB an.

Achtung: Das Blatt "2026" in der Objekt-1-Datei enthält den Zeitraum 2025.
"""

from __future__ import annotations

import os
import re
import sys
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from sqlalchemy import select

from app import models
from app.db import SessionLocal

# Dateipfad als Argument oder Umgebungsvariable (die Quelldatei ist nicht versioniert).
ULRICH_FILE = os.environ.get("IMPORT_XLSX") or (sys.argv[1] if len(sys.argv) > 1 else None)

WHG_RE = re.compile(r"^Whg\.?\s+(.+)")


def _num(v):
    if v is None or isinstance(v, str):
        return None
    return Decimal(str(v))


def parse_sheet(ws):
    """Extrahiert (period_start, period_end, tenants, garden) aus einem Jahres-Blatt."""
    period_start = period_end = None
    tenants: list[dict] = []
    garden: list[tuple[str, Decimal | None, Decimal | None]] = []
    cur = None

    for row in ws.iter_rows():
        vals = [c.value for c in row]
        b = vals[1] if len(vals) > 1 else None
        c = vals[2] if len(vals) > 2 else None
        d = vals[3] if len(vals) > 3 else None
        e = vals[4] if len(vals) > 4 else None

        if b == "Abrechnungszeitraum":
            period_start, period_end = c, d
            if isinstance(period_start, datetime):
                period_start = period_start.date()
            if isinstance(period_end, datetime):
                period_end = period_end.date()

        if isinstance(b, str):
            m = WHG_RE.match(b)
            if m:
                cur = {"name": m.group(1).strip(), "water": (None, None), "wm": (None, None), "area": None, "advance_year": None}
                tenants.append(cur)

        if cur is not None:
            if isinstance(c, str) and c == "Wasserzähler":
                cur["water"] = (_num(d), _num(e))
            elif isinstance(c, str) and c == "Waschmaschine":
                cur["wm"] = (_num(d), _num(e))
            elif isinstance(b, str) and b == "Grundsteuer" and cur["area"] is None:
                cur["area"] = _num(vals[5] if len(vals) > 5 else None)
            elif isinstance(b, str) and b == "Ihre Nebenkosten Vorauszahlungen" and cur["advance_year"] is None:
                cur["advance_year"] = _num(vals[7] if len(vals) > 7 else None)

        if isinstance(c, str) and c == "Zähler Nord":
            garden.append(("Garten Nord", _num(d), _num(e)))
        elif isinstance(c, str) and c == "Zähler Süd":
            garden.append(("Garten Süd", _num(d), _num(e)))

    return period_start, period_end, tenants, garden


def _ensure_unit(session, prop, designation, living_area):
    unit = session.scalar(
        select(models.LeaseUnit).where(
            models.LeaseUnit.property_id == prop.id,
            models.LeaseUnit.designation == designation,
        )
    )
    if unit is None:
        unit = models.LeaseUnit(property_id=prop.id, designation=designation, living_area=0, extra_area=0)
        session.add(unit)
        session.flush()
    if living_area is not None:
        unit.living_area = living_area
    return unit


def _ensure_meter(session, name, mtype, prop=None, unit=None):
    meter = session.scalar(select(models.Meter).where(models.Meter.name == name))
    if meter is None:
        meter = models.Meter(
            name=name, meter_type=mtype,
            property_id=prop.id if prop else None,
            lease_unit_id=unit.id if unit else None,
        )
        session.add(meter)
        session.flush()
    return meter


def _ensure_tenant(session, unit, name, move_in):
    tenant = session.scalar(
        select(models.Tenant).where(models.Tenant.lease_unit_id == unit.id, models.Tenant.name == name)
    )
    if tenant is None:
        tenant = models.Tenant(
            lease_unit_id=unit.id, name=name, move_in=move_in, move_out=None, monthly_advance=0
        )
        session.add(tenant)
        session.flush()
    elif tenant.move_in is None or move_in < tenant.move_in:
        tenant.move_in = move_in
    return tenant


def _add_reading(session, meter, rdate, value) -> int:
    if value is None or value <= 0:
        return 0
    exists = session.scalar(
        select(models.MeterReading).where(
            models.MeterReading.meter_id == meter.id,
            models.MeterReading.reading_date == rdate,
        )
    )
    if exists is None:
        session.add(models.MeterReading(meter_id=meter.id, reading_date=rdate, value=value))
        session.flush()  # autoflush ist in db.py deaktiviert → sofort sichtbar machen
        return 1
    return 0


def import_ulrich(session) -> int:
    wb = openpyxl.load_workbook(ULRICH_FILE, data_only=True)

    if not ULRICH_FILE:
        raise SystemExit("Bitte Excel-Dateipfad angeben (Argument oder Umgebungsvariable IMPORT_XLSX).")
    prop = session.scalar(select(models.Property).where(models.Property.name == "Objekt 1"))
    if prop is None:
        prop = models.Property(name="Objekt 1", street="", zip_code="", city="")
        session.add(prop)
        session.flush()

    # Sauberer Neustart der Wasser-/Einheitendaten dieses Objekts (idempotent).
    for unit in list(prop.lease_units):
        session.delete(unit)
    for meter in list(prop.meters):
        session.delete(meter)
    session.flush()

    # Gemeinschaftsfläche (separates Zimmer) – nur für den NF-Nenner.
    _ensure_unit(session, prop, "Gemeinschaftsfläche", Decimal("0")).extra_area = Decimal("12")

    added = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        period_start, period_end, tenants, garden = parse_sheet(ws)
        if not isinstance(period_start, date) or not isinstance(period_end, date):
            continue

        for gname, start, end in garden:
            meter = _ensure_meter(session, gname, models.MeterType.GARDEN, prop=prop)
            added += _add_reading(session, meter, period_start, start)
            added += _add_reading(session, meter, period_end, end)

        for t in tenants:
            name = t["name"]
            unit = _ensure_unit(session, prop, f"Wohnung {name}", t["area"])
            tenant = _ensure_tenant(session, unit, name, period_start)
            if t["advance_year"] is not None:
                tenant.monthly_advance = t["advance_year"] / Decimal("12")

            wm = _ensure_meter(session, f"{name} Wasser", models.MeterType.APARTMENT_WATER, unit=unit)
            wm_m = _ensure_meter(session, f"{name} Waschmaschine", models.MeterType.WASHING_MACHINE, unit=unit)

            added += _add_reading(session, wm, period_start, t["water"][0])
            added += _add_reading(session, wm, period_end, t["water"][1])
            added += _add_reading(session, wm_m, period_start, t["wm"][0])
            added += _add_reading(session, wm_m, period_end, t["wm"][1])

    session.commit()
    return added


if __name__ == "__main__":
    session = SessionLocal()
    try:
        count = import_ulrich(session)
        print(f"Import abgeschlossen. {count} neue Zählerstände eingespielt.")
        meters = session.scalars(select(models.Meter).order_by(models.Meter.name)).all()
        for m in meters:
            readings = len(m.readings)
            if readings:
                print(f"  {m.name:30s} {readings:3d} Stände")
    finally:
        session.close()
